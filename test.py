import os
from dotenv import load_dotenv
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
import openpyxl

load_dotenv()

sg_api_key = os.getenv('SENDGRID_API_KEY')
sg = SendGridAPIClient(sg_api_key)

archivo_excel = openpyxl.load_workbook("direcciones_correo.xlsx")
hoja_activa = archivo_excel.active

html = open('email.html', 'r')
contenido_html = html.read()

for fila in range(2, hoja_activa.max_row + 1):
    destinatario = hoja_activa.cell(row=fila, column=1).value

    mensaje = Mail(
        from_email='lukijian@gmail.com',
        to_emails=destinatario,
        subject='Sus Credenciales estan bajo ataque',
        html_content=contenido_html
    )

    try:
        respuesta = sg.send(mensaje)
        print(f"Correo enviado a {destinatario}")
        print(respuesta.status_code)
        print(respuesta.body)
        print(respuesta.headers)
    except Exception as e:
        print(f"Error al enviar el correo a {destinatario}: {e.message}")

print("Proceso finalizado")
